import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# ==================================================
# PATHS
# ==================================================
historical_file = Path("data/raw/global/ACLED.csv")
region_dir = Path("data/raw/global/regions_2026")

output_dir = Path("data/cleaned/global")
output_dir.mkdir(parents=True, exist_ok=True)

output_main = output_dir / "conflict_standardized_monthlybytype.csv"
output_country = output_dir / "conflict_country_monthlybytype.csv"

# ==================================================
# HELPERS
# ==================================================
def read_csv_or_excel(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path)
    return pd.read_excel(path, engine="openpyxl")


def read_excel_robust(path: Path) -> pd.DataFrame:
    try:
        return pd.read_excel(path, engine="openpyxl")
    except Exception as e:
        print(f"[WARN] pandas failed for {path.name}: {e}")
        print("[INFO] Trying openpyxl fallback...")

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.values)

        if not rows:
            raise ValueError(f"{path.name} appears empty.")

        header = list(rows[0])
        data = rows[1:]
        return pd.DataFrame(data, columns=header)


def clean_text_cols(df: pd.DataFrame, cols):
    for col in cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})
    return df


def build_month_fields(df: pd.DataFrame, date_col: str):
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    df = df[df[date_col].notna()].copy()
    df["year"] = df[date_col].dt.year
    df["month_num"] = df[date_col].dt.month
    df["month"] = df[date_col].dt.strftime("%B")
    return df


# ==================================================
# 1) HISTORICAL EVENT-LEVEL ACLED (through 2025)
# ==================================================
hist = read_csv_or_excel(historical_file)

rename_map_hist = {
    "iso": "iso3",
    "country": "country",
    "admin1": "admin1",
    "admin2": "admin2",
    "event_date": "event_date",
    "event_type": "event_type",
    "fatalities": "fatalities",
}

hist = hist.rename(columns=rename_map_hist)

needed_hist = ["iso3", "country", "admin1", "admin2", "event_date", "event_type", "fatalities"]
missing_hist = [c for c in needed_hist if c not in hist.columns]
if missing_hist:
    raise ValueError(f"Missing expected historical columns: {missing_hist}")

hist = hist[needed_hist].copy()
hist = clean_text_cols(hist, ["iso3", "country", "admin1", "admin2", "event_type"])
hist = build_month_fields(hist, "event_date")

hist["fatalities"] = pd.to_numeric(hist["fatalities"], errors="coerce").fillna(0)
hist["iso3"] = pd.to_numeric(hist["iso3"], errors="coerce")

hist = hist[
    hist["country"].notna()
    & hist["iso3"].notna()
    & hist["event_type"].notna()
].copy()

hist["iso3"] = hist["iso3"].astype(int)
hist = hist[hist["year"] <= 2025].copy()

# detailed monthly admin table by event type
main_by_type_hist = (
    hist.groupby(
        ["iso3", "country", "admin1", "admin2", "year", "month_num", "month", "event_type"],
        dropna=False,
        as_index=False,
    )
    .agg(
        events=("event_date", "size"),
        fatalities=("fatalities", "sum"),
    )
)

# detailed monthly admin table for all event types
main_all_hist = (
    hist.groupby(
        ["iso3", "country", "admin1", "admin2", "year", "month_num", "month"],
        dropna=False,
        as_index=False,
    )
    .agg(
        events=("event_date", "size"),
        fatalities=("fatalities", "sum"),
    )
)

main_all_hist["event_type"] = "All"
main_hist = pd.concat([main_all_hist, main_by_type_hist], ignore_index=True)

# historical country monthly by event type
country_by_type_hist = (
    hist.groupby(
        ["iso3", "country", "year", "month_num", "month", "event_type"],
        as_index=False,
    )
    .agg(
        events=("event_date", "size"),
        fatalities=("fatalities", "sum"),
    )
)

# historical country monthly for all event types
country_all_hist = (
    hist.groupby(
        ["iso3", "country", "year", "month_num", "month"],
        as_index=False,
    )
    .agg(
        events=("event_date", "size"),
        fatalities=("fatalities", "sum"),
    )
)

country_all_hist["event_type"] = "All"
country_hist = pd.concat([country_all_hist, country_by_type_hist], ignore_index=True)

# ==================================================
# 2) READ ALL REGIONAL FILES (2026 only)
# ==================================================
if not region_dir.exists():
    raise ValueError(f"Folder does not exist: {region_dir}")

regional_files = sorted(region_dir.glob("*.xlsx"))
if not regional_files:
    raise ValueError(f"No .xlsx files found in: {region_dir}")

frames = []
for file in regional_files:
    print(f"Reading: {file.name}")
    temp = read_excel_robust(file)
    temp["source_file"] = file.name
    frames.append(temp)

regions = pd.concat(frames, ignore_index=True)

rename_map_regions = {
    "WEEK": "event_date",
    "REGION": "region",
    "COUNTRY": "country",
    "ADMIN1": "admin1",
    "EVENT_TYPE": "event_type",
    "SUB_EVENT_TYPE": "sub_event_type",
    "EVENTS": "events",
    "FATALITIES": "fatalities",
    "POPULATION_EXPOSURE": "population_exposure",
    "DISORDER_TYPE": "disorder_type",
    "ID": "record_id",
    "CENTROID_LATITUDE": "centroid_latitude",
    "CENTROID_LONGITUDE": "centroid_longitude",
}

regions = regions.rename(columns=rename_map_regions)

needed_regions = [
    "event_date",
    "region",
    "country",
    "admin1",
    "event_type",
    "events",
    "fatalities",
    "population_exposure",
    "centroid_latitude",
    "centroid_longitude",
]

missing_regions = [c for c in needed_regions if c not in regions.columns]
if missing_regions:
    raise ValueError(f"Missing expected regional columns: {missing_regions}")

regions = regions[needed_regions].copy()
regions = clean_text_cols(regions, ["region", "country", "admin1", "event_type"])
regions = build_month_fields(regions, "event_date")

regions["events"] = pd.to_numeric(regions["events"], errors="coerce").fillna(0)
regions["fatalities"] = pd.to_numeric(regions["fatalities"], errors="coerce").fillna(0)
regions["population_exposure"] = pd.to_numeric(regions["population_exposure"], errors="coerce").fillna(0)
regions["centroid_latitude"] = pd.to_numeric(regions["centroid_latitude"], errors="coerce")
regions["centroid_longitude"] = pd.to_numeric(regions["centroid_longitude"], errors="coerce")

regions = regions[
    regions["country"].notna()
    & regions["event_type"].notna()
].copy()

regions = regions[regions["year"] == 2026].copy()

# ==================================================
# 3) MAP COUNTRY -> ISO NUMERIC
# ==================================================
country_iso_lookup = (
    hist[["country", "iso3"]]
    .dropna()
    .drop_duplicates()
    .groupby("country", as_index=False)["iso3"]
    .first()
)

regions = regions.merge(country_iso_lookup, how="left", on="country")

manual_iso_fallback = {
    "Anguilla": 660,
    "Bailiwick of Guernsey": 831,
    "Barbados": 52,
    "British Indian Ocean Territory": 86,
    "British Virgin Islands": 92,
    "Falkland Islands": 238,
    "Faroe Islands": 234,
    "Gibraltar": 292,
    "Isle of Man": 833,
    "Liechtenstein": 438,
    "Montserrat": 500,
    "Qatar": 634,
    "Saint Helena, Ascension and Tristan da Cunha": 654,
    "Saint Kitts and Nevis": 659,
    "Saint Vincent and the Grenadines": 670,
    "San Marino": 674,
    "Turks and Caicos Islands": 796,
    "Vatican City": 336,
    "Virgin Islands, U.S.": 850,
}

regions["iso3"] = regions["iso3"].fillna(regions["country"].map(manual_iso_fallback))

missing_iso_countries = sorted(
    regions.loc[regions["iso3"].isna(), "country"].dropna().unique().tolist()
)

if missing_iso_countries:
    print("Still missing ISO mapping:")
    print(missing_iso_countries)
    raise ValueError("Some countries are still missing ISO numeric codes after fallback mapping.")

regions["iso3"] = regions["iso3"].astype(int)

# ==================================================
# 4) 2026 COUNTRY MONTHLY
# ==================================================
country_by_type_2026 = (
    regions.groupby(
        ["iso3", "country", "year", "month_num", "month", "event_type"],
        as_index=False,
    )
    .agg(
        events=("events", "sum"),
        fatalities=("fatalities", "sum"),
    )
)

country_all_2026 = (
    regions.groupby(
        ["iso3", "country", "year", "month_num", "month"],
        as_index=False,
    )
    .agg(
        events=("events", "sum"),
        fatalities=("fatalities", "sum"),
    )
)

country_all_2026["event_type"] = "All"
country_2026 = pd.concat([country_all_2026, country_by_type_2026], ignore_index=True)

# ==================================================
# 5) 2026 ADMIN1 MONTHLY
# ==================================================
main_by_type_2026 = (
    regions.groupby(
        ["iso3", "country", "admin1", "year", "month_num", "month", "event_type"],
        as_index=False,
    )
    .agg(
        events=("events", "sum"),
        fatalities=("fatalities", "sum"),
        population_exposure=("population_exposure", "sum"),
    )
)

main_by_type_2026["admin2"] = pd.NA

main_all_2026 = (
    regions.groupby(
        ["iso3", "country", "admin1", "year", "month_num", "month"],
        as_index=False,
    )
    .agg(
        events=("events", "sum"),
        fatalities=("fatalities", "sum"),
        population_exposure=("population_exposure", "sum"),
    )
)

main_all_2026["admin2"] = pd.NA
main_all_2026["event_type"] = "All"

main_2026 = pd.concat([main_all_2026, main_by_type_2026], ignore_index=True)

# align columns
if "population_exposure" not in main_hist.columns:
    main_hist["population_exposure"] = pd.NA

main_hist = main_hist[
    ["iso3", "country", "admin1", "admin2", "year", "month_num", "month", "event_type", "events", "fatalities", "population_exposure"]
]
main_2026 = main_2026[
    ["iso3", "country", "admin1", "admin2", "year", "month_num", "month", "event_type", "events", "fatalities", "population_exposure"]
]

main = pd.concat([main_hist, main_2026], ignore_index=True)
country_monthly = pd.concat([country_hist, country_2026], ignore_index=True)

# ==================================================
# 6) SORT + SAVE
# ==================================================
main = main.sort_values(
    ["year", "month_num", "country", "admin1", "admin2", "event_type"]
).reset_index(drop=True)

country_monthly = country_monthly.sort_values(
    ["year", "month_num", "country", "event_type"]
).reset_index(drop=True)

main.to_csv(output_main, index=False, encoding="utf-8-sig")
country_monthly.to_csv(output_country, index=False, encoding="utf-8-sig")

print("Saved:")
print(output_main)
print(output_country)

print("\nYears in main:")
print(sorted(main["year"].dropna().unique().tolist()))

print("\nYears in country:")
print(sorted(country_monthly["year"].dropna().unique().tolist()))